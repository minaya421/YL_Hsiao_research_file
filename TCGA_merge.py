# -*- coding: utf-8 -*-
"""
Created on Tue Aug  6 16:52:10 2024

@author: minay
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

case_id_path = '合併檔案一路徑'
case_id_df = pd.read_excel(case_id_path)

case_data_path ='合併檔案二路徑'
case_data_df = pd.read_excel(case_data_path)

merged_df = case_data_df[case_data_df['case_id'].isin(case_id_df['case_id'])]

output_file_path = '合併後輸出檔案路徑'
merged_df.to_excel(output_file_path, index=False)

wb = load_workbook(output_file_path)
ws = wb.active
for col in ws.columns:
    length = max(len(str(cell.value))if cell.value is not None else 0 for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = length + 6
wb.save(output_file_path)

print('Finish')
